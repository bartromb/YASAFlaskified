"""
generate_excel_report.py — Multi-sheet Excel export voor YASAFlaskified
Gebruik: xlsx_path = generate_excel_report(results_dict, output_path)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── Kleuren (hex zonder #) ──
BG_BLAUW   = "1A3A5C"
BG_LICHT   = "D6E4F0"
BG_GRIJS   = "ECF0F1"
TXT_WIT    = "FFFFFF"
TXT_BLAUW  = "1A3A5C"
TXT_GROEN  = "27AE60"
TXT_ROOD   = "C0392B"
TXT_ORANJE = "E67E22"

STAGE_BG = {
    "W":  "E74C3C",
    "N1": "F39C12",
    "N2": "3498DB",
    "N3": "2C3E50",
    "R":  "9B59B6",
}


# ─────────────────────────────────────────────
# STIJL HELPERS
# ─────────────────────────────────────────────

def hdr_font(size=10, bold=True, color=TXT_WIT):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=9, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def hdr_fill(color=BG_BLAUW):
    return PatternFill("solid", fgColor=color)

def alt_fill(row_idx, even_color=BG_GRIJS):
    return PatternFill("solid", fgColor=even_color) if row_idx % 2 == 0 else PatternFill()

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header_row(ws, row, headers, start_col=1, bg=BG_BLAUW, span=None):
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = hdr_font(color=TXT_WIT)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = center()
        c.border = thin_border()

def write_data_row(ws, row, values, start_col=1, bold=False):
    for i, v in enumerate(values, start=start_col):
        c = ws.cell(row=row, column=i, value=v)
        c.font = body_font(bold=bold)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()
        if row % 2 == 0:
            c.fill = PatternFill("solid", fgColor=BG_GRIJS)

def col_widths(ws, widths: dict):
    """widths = {'A': 20, 'B': 15, ...}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def sheet_title(ws, title: str, subtitle: str = "", span: int = 6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Arial", size=14, bold=True, color=TXT_WIT)
    c.fill = PatternFill("solid", fgColor=BG_BLAUW)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(name="Arial", size=9, color="7F8C8D")
        c2.fill = PatternFill("solid", fgColor=BG_LICHT)
        c2.alignment = center()
        ws.row_dimensions[2].height = 18

    return 4  # eerste beschikbare datarij


def safe(v, decimals=2):
    """Veilig omzetten naar float of str voor Excel."""
    if v is None:
        return "—"
    try:
        f = float(v)
        return round(f, decimals)
    except (TypeError, ValueError):
        return str(v)


# ─────────────────────────────────────────────
# SHEET 1: OVERZICHT
# ─────────────────────────────────────────────

def build_overview(wb, results):
    ws = wb.active
    ws.title = "Overzicht"
    meta = results.get("meta", {})
    stats = results.get("sleep_statistics", {}).get("stats", {})

    start = sheet_title(ws, "SleepAI — Slaapanalyse Overzicht",
                        f"EEG: {meta.get('eeg_channel','—')}  |  Duur: {meta.get('duration_min','—')} min  |  YASA v{meta.get('yasa_version','—')}",
                        span=4)
    ws.row_dimensions[start - 1].height = 6

    # KPI blok
    kpis = [
        ("Totale slaaptijd (TST)", safe(stats.get("TST")), "min"),
        ("Slaapefficiëntie (SE)",  safe(stats.get("SE")),  "%"),
        ("Slaapladentie (SOL)",   safe(stats.get("SOL")), "min"),
        ("WASO",                   safe(stats.get("WASO")),"min"),
        ("N1 %",  safe(stats.get("N1")),  "%"),
        ("N2 %",  safe(stats.get("N2")),  "%"),
        ("N3 %",  safe(stats.get("N3")),  "%"),
        ("REM %", safe(stats.get("REM")), "%"),
    ]
    write_header_row(ws, start, ["Indicator", "Waarde", "Eenheid", ""], span=4)
    for i, (label, val, unit) in enumerate(kpis, start=start+1):
        write_data_row(ws, i, [label, val, unit, ""])
        ws.cell(row=i, column=2).font = Font(name="Arial", size=11, bold=True, color=TXT_BLAUW)
        ws.cell(row=i, column=2).alignment = center()

    col_widths(ws, {"A": 28, "B": 14, "C": 10, "D": 20})

    # Spindles & SW samenvatting
    after_kpi = start + len(kpis) + 2
    ws.cell(row=after_kpi, column=1, value="Spindles").font = Font(bold=True, color=TXT_BLAUW, size=10)
    sp = results.get("spindles", {})
    ws.cell(row=after_kpi+1, column=1, value="Totaal gedetecteerd")
    ws.cell(row=after_kpi+1, column=2, value=sp.get("total_spindles", "—"))

    sw = results.get("slow_waves", {})
    ws.cell(row=after_kpi+3, column=1, value="Trage golven (N3)").font = Font(bold=True, color=TXT_BLAUW, size=10)
    ws.cell(row=after_kpi+4, column=1, value="Totaal gedetecteerd")
    ws.cell(row=after_kpi+4, column=2, value=sw.get("total_slow_waves", "—"))

    rem = results.get("rem", {}).get("summary", {})
    ws.cell(row=after_kpi+6, column=1, value="REM").font = Font(bold=True, color=TXT_BLAUW, size=10)
    ws.cell(row=after_kpi+7, column=1, value="REM perioden")
    ws.cell(row=after_kpi+7, column=2, value=rem.get("n_rem_periods", "—"))
    ws.cell(row=after_kpi+8, column=1, value="Totale REM (min)")
    ws.cell(row=after_kpi+8, column=2, value=safe(rem.get("rem_duration_min")))

    art = results.get("artifacts", {}).get("summary", {})
    ws.cell(row=after_kpi+10, column=1, value="Artefacten").font = Font(bold=True, color=TXT_BLAUW, size=10)
    ws.cell(row=after_kpi+11, column=1, value="Artefact epochs")
    ws.cell(row=after_kpi+11, column=2, value=art.get("n_artifact_epochs", "—"))
    ws.cell(row=after_kpi+12, column=1, value="Artefact %")
    ws.cell(row=after_kpi+12, column=2, value=safe(art.get("artifact_percent")))
    pct_cell = ws.cell(row=after_kpi+12, column=2)
    pct = art.get("artifact_percent", 0) or 0
    pct_cell.font = Font(name="Arial", size=9, bold=True,
                         color=TXT_ROOD if pct > 20 else (TXT_ORANJE if pct > 10 else TXT_GROEN))


# ─────────────────────────────────────────────
# SHEET 2: ALLE STATISTIEKEN
# ─────────────────────────────────────────────

def build_stats_sheet(wb, results):
    ws = wb.create_sheet("Slaapstatistieken")
    stats = results.get("sleep_statistics", {}).get("stats", {})
    start = sheet_title(ws, "Slaapstatistieken — AASM Normen", span=3)
    write_header_row(ws, start, ["Parameter", "Waarde", "Eenheid"], span=3)
    r = start + 1
    for k, v in stats.items():
        unit = "min" if any(x in k for x in ["TST","SOL","WASO","Lat","TRT"]) else (
               "%" if any(x in k for x in ["SE","N1","N2","N3","REM","W"]) else "")
        write_data_row(ws, r, [k, safe(v), unit])
        r += 1
    col_widths(ws, {"A": 30, "B": 14, "C": 10})


# ─────────────────────────────────────────────
# SHEET 3: HYPNOGRAM
# ─────────────────────────────────────────────

def build_hypno_sheet(wb, results):
    ws = wb.create_sheet("Hypnogram")
    timeline = results.get("hypnogram_timeline", {}).get("timeline", [])
    start = sheet_title(ws, "Hypnogram Tijdlijn", span=4)
    write_header_row(ws, start, ["Epoch", "Tijd", "Slaapfase", "Tijdstip (min)"], span=4)
    for i, ep in enumerate(timeline, start=start+1):
        stage = ep.get("stage", "W")
        write_data_row(ws, i, [ep.get("epoch"), ep.get("time",""), stage, safe(ep.get("time_min"))])
        bg = STAGE_BG.get(stage, "AAAAAA")
        ws.cell(row=i, column=3).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=i, column=3).font = Font(name="Arial", size=9, bold=True, color=TXT_WIT)
        ws.cell(row=i, column=3).alignment = center()
    col_widths(ws, {"A": 8, "B": 12, "C": 12, "D": 16})


# ─────────────────────────────────────────────
# SHEET 4: SPINDELS
# ─────────────────────────────────────────────

def build_spindles_sheet(wb, results):
    ws = wb.create_sheet("Spindels")
    sp = results.get("spindles", {})
    start = sheet_title(ws, f"Spindle Detectie — {sp.get('total_spindles',0)} gedetecteerd", span=8)

    summary = sp.get("summary", [])
    if summary:
        ws.cell(row=start, column=1, value="Per-kanaal samenvatting").font = Font(bold=True, color=TXT_BLAUW)
        keys = list(summary[0].keys())
        write_header_row(ws, start+1, keys, span=len(keys))
        for i, row in enumerate(summary, start=start+2):
            write_data_row(ws, i, [safe(row.get(k)) for k in keys])
        start = start + len(summary) + 4

    events = sp.get("spindles", [])
    if events:
        ws.cell(row=start, column=1, value="Per-event tabel").font = Font(bold=True, color=TXT_BLAUW)
        keys = list(events[0].keys())
        write_header_row(ws, start+1, keys, span=len(keys))
        for i, ev in enumerate(events, start=start+2):
            write_data_row(ws, i, [safe(ev.get(k)) for k in keys])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 14


# ─────────────────────────────────────────────
# SHEET 5: TRAGE GOLVEN
# ─────────────────────────────────────────────

def build_sw_sheet(wb, results):
    ws = wb.create_sheet("Trage Golven")
    sw = results.get("slow_waves", {})
    start = sheet_title(ws, f"Slow-Wave Detectie — {sw.get('total_slow_waves',0)} gedetecteerd (N3)", span=8)

    summary = sw.get("summary", [])
    if summary:
        keys = list(summary[0].keys())
        write_header_row(ws, start, keys, span=len(keys))
        for i, row in enumerate(summary, start=start+1):
            write_data_row(ws, i, [safe(row.get(k)) for k in keys])
        start = start + len(summary) + 3

    events = sw.get("slow_waves", [])
    if events:
        ws.cell(row=start, column=1, value="Per-event tabel").font = Font(bold=True, color=TXT_BLAUW)
        keys = list(events[0].keys())
        write_header_row(ws, start+1, keys, span=len(keys))
        for i, ev in enumerate(events, start=start+2):
            write_data_row(ws, i, [safe(ev.get(k)) for k in keys])

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 14


# ─────────────────────────────────────────────
# SHEET 6: REM
# ─────────────────────────────────────────────

def build_rem_sheet(wb, results):
    ws = wb.create_sheet("REM")
    rem = results.get("rem", {})
    start = sheet_title(ws, "REM Detectie & Transities", span=4)

    rs = rem.get("summary", {})
    write_header_row(ws, start, ["Maat", "Waarde"], span=2)
    r = start + 1
    for k, v in rs.items():
        write_data_row(ws, r, [k, safe(v)])
        r += 1

    transitions = rem.get("transitions", [])
    if transitions:
        r += 2
        ws.cell(row=r, column=1, value="NREM → REM Transities").font = Font(bold=True, color=TXT_BLAUW)
        r += 1
        write_header_row(ws, r, ["#", "Epoch", "Van fase", "Tijdstip (min)"], span=4)
        for i, t in enumerate(transitions, start=r+1):
            write_data_row(ws, i, [i - r, t["epoch"], t["from_stage"], safe(t["to_REM_min"])])

    col_widths(ws, {"A": 28, "B": 16, "C": 14, "D": 18})


# ─────────────────────────────────────────────
# SHEET 7: BANDVERMOGEN
# ─────────────────────────────────────────────

def build_bandpower_sheet(wb, results):
    ws = wb.create_sheet("Bandvermogen")
    bp = results.get("bandpower", {})
    bands = ["delta", "theta", "alpha", "sigma", "beta", "gamma"]
    per_stage = bp.get("per_stage", {})
    start = sheet_title(ws, "Bandvermogen — Relatief per slaapfase", span=len(bands)+1)

    write_header_row(ws, start, ["Fase"] + [b.capitalize() for b in bands], span=len(bands)+1)
    r = start + 1
    for stage, bdata in per_stage.items():
        row_vals = [stage] + [safe(bdata.get(b), 4) for b in bands]
        write_data_row(ws, r, row_vals)
        bg = STAGE_BG.get(stage, "AAAAAA")
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=1).font = Font(name="Arial", size=9, bold=True, color=TXT_WIT)
        ws.cell(row=r, column=1).alignment = center()
        r += 1

    # Band ratio's
    ratios = bp.get("band_ratios", {})
    if ratios:
        r += 2
        ws.cell(row=r, column=1, value="Band Ratio's").font = Font(bold=True, color=TXT_BLAUW)
        r += 1
        write_header_row(ws, r, ["Ratio", "Waarde"], span=2)
        for ratio_name, ratio_val in ratios.items():
            r += 1
            write_data_row(ws, r, [ratio_name, safe(ratio_val, 3)])

    # Radar chart
    if per_stage:
        chart_start_row = start + 1
        n_stages = len(per_stage)
        chart = RadarChart()
        chart.type = "filled"
        chart.style = 10
        chart.title = "Bandvermogen per slaapfase"
        chart.y_axis.numFmt = "0.00"
        chart.shape = 4

        data_ref = Reference(ws,
                             min_col=2, max_col=len(bands)+1,
                             min_row=start, max_row=start + n_stages)
        cats = Reference(ws, min_col=2, max_col=len(bands)+1, min_row=start)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=chart_start_row,
                                       max_row=chart_start_row + n_stages - 1))
        chart.width = 15
        chart.height = 12
        ws.add_chart(chart, f"I{start}")

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 13


# ─────────────────────────────────────────────
# SHEET 8: CYCLI
# ─────────────────────────────────────────────

def build_cycles_sheet(wb, results):
    ws = wb.create_sheet("Slaapcycli")
    cyc = results.get("sleep_cycles", {})
    cycles = cyc.get("cycles", [])
    start = sheet_title(ws, f"Slaapcycli — {cyc.get('n_cycles',0)} cycli gedetecteerd", span=5)
    write_header_row(ws, start, ["Cyclus", "Start epoch", "Eind epoch", "Duur (min)", "Stage-samenstelling"], span=5)
    for i, c in enumerate(cycles, start=start+1):
        dist = " | ".join(f"{s}: {p}%" for s, p in c.get("stage_distribution", {}).items())
        write_data_row(ws, i, [c["cycle"], c["start_epoch"], c["end_epoch"], safe(c["duration_min"]), dist])
    col_widths(ws, {"A": 10, "B": 14, "C": 14, "D": 12, "E": 50})


# ─────────────────────────────────────────────
# SHEET 9: ARTEFACTEN
# ─────────────────────────────────────────────

def build_artifacts_sheet(wb, results):
    ws = wb.create_sheet("Artefacten")
    art = results.get("artifacts", {})
    s = art.get("summary", {})
    start = sheet_title(ws,
        f"Artefactdetectie — {s.get('n_artifact_epochs',0)} / {s.get('n_total_epochs',0)} epochs ({s.get('artifact_percent',0)}%)",
        span=4)
    write_header_row(ws, start, ["Epoch", "Max amplitude (μV)", "Vlat signaal", "Hoge amplitude"], span=4)
    for i, ep in enumerate(art.get("artifact_epochs", []), start=start+1):
        write_data_row(ws, i, [
            ep.get("epoch"),
            safe(ep.get("max_amplitude_uV")),
            "JA" if ep.get("flat_signal") else "",
            "JA" if ep.get("high_amplitude") else "",
        ])
        if ep.get("high_amplitude"):
            ws.cell(row=i, column=4).font = Font(name="Arial", size=9, bold=True, color=TXT_ROOD)
    col_widths(ws, {"A": 10, "B": 22, "C": 16, "D": 16})


# ─────────────────────────────────────────────
# HOOFD GENERATOR
# ─────────────────────────────────────────────

def generate_excel_report(results: dict, output_path: str) -> str:
    """
    Genereer een volledig Excel-werkboek met alle analyseresultaten.

    Parameters
    ----------
    results     : dict  — uitvoer van run_full_analysis()
    output_path : str   — pad waar het .xlsx bestand opgeslagen wordt

    Returns
    -------
    str — output_path
    """
    wb = Workbook()

    build_overview(wb, results)
    build_stats_sheet(wb, results)
    build_hypno_sheet(wb, results)
    build_spindles_sheet(wb, results)
    build_sw_sheet(wb, results)
    build_rem_sheet(wb, results)
    build_bandpower_sheet(wb, results)
    build_cycles_sheet(wb, results)
    build_artifacts_sheet(wb, results)

    wb.save(output_path)
    return output_path
